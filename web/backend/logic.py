"""
NoNA Core Logic
Refactored from NoNA.py for usage in Web Backend.
Removes Rhino dependencies.
"""

from typing import Dict, List, Any, Tuple
import math
import csv
import io
from typing import Dict, List, Any, Tuple

# ==============================================================================
# CONFIGURATION & CONSTANTS
# ==============================================================================
# Defaults for fallback
DEFAULT_PARAMS = {
    'COST_DEMOLITION_M2': 1600.0,
    'COST_LICENSE_M2': 15.0,
    'COST_WASTE_PERCENT': 0.15,
    'PARKING_M2_PER_SPOT': 12.5,
    'PARKING_DRIVEWAY_FACTOR': 1.50,
    # Indirects
    'PCT_HONORARIOS': 15.0,
    'PCT_LEGALES': 2.0,
    'PCT_ADM': 10.0,
    'PCT_FIN': 3.0,
    'PCT_COM': 6.0
}

CONSTANTS = {
    'PARKING_FACTORS': {
        'centro': {'comercial': 35},
        'poniente': {'comercial': 25},
        'norte': {'comercial': 30},
        'sur': {'comercial': 25}
    }
}

def calculate_land_metrics(area_terreno: float, cost_per_unit: float) -> Tuple[float, float, str, str]:
    """Calculates total land value from numeric area input."""
    total_value = area_terreno * cost_per_unit
    return area_terreno, total_value, f"{area_terreno:.2f} m2", f"${total_value:,.2f} mxn"

def calculate_demolition_cost(
    do_demolition: bool, 
    area_demolition: float, 
    land_area: float,
    params: Dict[str, float]
) -> Tuple[float, str]:
    """Computes demolition, license, and waste removal costs."""
    if not do_demolition:
        return 0.0, "$0.00 mxn"
        
    cost_dem = area_demolition * params.get('COST_DEMOLITION_M2', DEFAULT_PARAMS['COST_DEMOLITION_M2'])
    cost_lic = land_area * params.get('COST_LICENSE_M2', DEFAULT_PARAMS['COST_LICENSE_M2'])
    cost_res = land_area * params.get('COST_WASTE_PERCENT', DEFAULT_PARAMS['COST_WASTE_PERCENT'])
    
    total = cost_dem + cost_lic + cost_res
    return total, f"${total:,.2f} mxn"

def calculate_regulatory_areas(
    land_area: float, 
    cos: float, 
    cus: float, 
    cas: float, 
    retiros_area: float
) -> Dict[str, float]:
    """Calculates permitted construction areas based on coefficients."""
    return {
        'cos_area': land_area * cos,
        'cus_area': land_area * cus,
        'cas_area': land_area * cas,
        'net_area': land_area - retiros_area
    }

def calculate_mixed_use(
    is_mixed: bool, 
    num_locales: int, 
    cos_area: float, 
    price_per_m2: float
) -> Dict[str, float]:
    """Calculates commercial area metrics if mixed-use is enabled."""
    results = {
        'area_local': 0.0,
        'venta_local': 0.0,
        'ingreso_total': 0.0,
        'area_comercio': 0.0
    }
    
    if not is_mixed:
        return results
        
    if num_locales > 0:
        # Logic from original: Commercial Area = COS Area
        results['area_comercio'] = cos_area 
        results['area_local'] = cos_area / float(num_locales)
        results['venta_local'] = results['area_local'] * price_per_m2
        results['ingreso_total'] = results['venta_local'] * num_locales
    
    return results

def calculate_parking(
    enable: bool, 
    n_viviendas: int, 
    cos_area: float, 
    circ_area: float,
    delegaciones: List[str],
    factors: List[float],
    cost_per_m2: float,
    params: Dict[str, float]
) -> Dict[str, Any]:
    """Calculates parking spots, area, and cost based on district."""
    if not enable:
        return {'cost': 0.0, 'area': 0.0, 'details': {}}
        
    cleaned_del = [d.strip().lower() for d in delegaciones]
    # NOTE: Logic assumes lists match, simplified for brevity
    
    total_cost = 0.0
    total_area_m2 = 0.0
    
    c_vivienda_list = []
    c_comercio_list = []
    c_total_list = []
    
    m2_spot = params.get('PARKING_M2_PER_SPOT', DEFAULT_PARAMS['PARKING_M2_PER_SPOT'])
    drive_factor = params.get('PARKING_DRIVEWAY_FACTOR', DEFAULT_PARAMS['PARKING_DRIVEWAY_FACTOR'])

    for dep, fac in zip(cleaned_del, factors):
        rules = CONSTANTS['PARKING_FACTORS'].get(dep)
        if not rules: continue
            
        c_viv = n_viviendas * fac
        c_com = (cos_area - circ_area) / rules['comercial'] if cos_area else 0
        
        spots = math.ceil(c_viv + c_com)
        area = spots * m2_spot * drive_factor
        cost = area * cost_per_m2
        
        total_cost += cost
        total_area_m2 += area
        
        c_vivienda_list.append(c_viv)
        c_comercio_list.append(c_com)
        c_total_list.append(spots)

    return {
        'cost': total_cost,
        'area': total_area_m2,
        'cost_per_m2': cost_per_m2,
        'details': {
            'cajones_vivienda': c_vivienda_list,
            'cajones_comercio': c_comercio_list,
            'cajones_total': c_total_list
        }
    }

def solve_target_price(
    cost_total: float, 
    desired_margin_percent: float,
    current_income: float
) -> Tuple[float, float, float]:
    """Deprecated: Replaced by inline advanced simulation accounting for variable rates."""
    return current_income, 0.0, 0.0

def run_calculation(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Main entry point for calculation.
    Expects a dictionary with all required inputs.
    """
    try:
        # Unpack inputs
        area_terreno = float(data.get('area_terreno', 0))
        valor_terreno = float(data.get('valor_terreno', 0))
        
        # Normativa
        COS = float(data.get('COS', 0))
        CUS = float(data.get('CUS', 0))
        CAS = float(data.get('CAS', 0))
        area_retiros = float(data.get('area_retiros', 0))
        
        # Demolition
        demolicion = bool(data.get('demolicion', False))
        area_demolicion = float(data.get('area_demolicion', 0))
        
        # Project
        n_viviendas = int(data.get('n_viviendas', 0))
        usos_mixtos = bool(data.get('usos_mixtos', False))
        num_locales = int(data.get('num_locales', 0))
        costo_local_m2 = float(data.get('costo_local_m2', 0))
        
        # Construction
        costoMetroConstruccion = float(data.get('costoMetroConstruccion', 0))
        Costo_de_venta_m2 = float(data.get('Costo_de_venta_m2', 0))
        areaCirculacionPorcentaje = float(data.get('areaCirculacionPorcentaje', 0))
        
        # Parking
        estacionamiento = bool(data.get('estacionamiento', False))
        tipo_estacionamiento = float(data.get('tipo_estacionamiento', 0))
        # Handle lists for parking
        delegacion = data.get('delegacion', [])
        if isinstance(delegacion, str): delegacion = [delegacion]
        
        Distrito = data.get('Distrito', [])
        if isinstance(Distrito, (int, float)): Distrito = [Distrito]
        
        # Simulation
        utilidadDeseada = float(data.get('utilidadDeseada', 20.0))
        correrSimulacion = bool(data.get('correrSimulacion', False))

        
        # Extract params from input data (passed from main.py)
        # We default to empty dict which will trigger DEFAULT_PARAMS in helpers
        params = data.get('parameters', {})
        
        # --- CALCULATION STEPS ---

        # 1. Land
        area_val, total_val, txt_area, txt_val = calculate_land_metrics(area_terreno, valor_terreno)

        # 2. Normative
        reg = calculate_regulatory_areas(area_val, COS, CUS, CAS, area_retiros)

        # 3. Demolition
        # Expanded breakdown
        dem_cost_only = 0.0
        lic_cost = 0.0
        res_cost = 0.0
        dem_txt = "$0.00 mxn"
        dem_cost = 0.0

        if demolicion:
             dem_cost_only = area_demolicion * params.get('COST_DEMOLITION_M2', DEFAULT_PARAMS['COST_DEMOLITION_M2'])
             lic_cost = area_val * params.get('COST_LICENSE_M2', DEFAULT_PARAMS['COST_LICENSE_M2'])
             # EXPERT FIX: Waste removal cost is a % of Demolition pure cost, not a % of total land area
             res_cost = dem_cost_only * params.get('COST_WASTE_PERCENT', DEFAULT_PARAMS['COST_WASTE_PERCENT']) 
             
             dem_cost = dem_cost_only + lic_cost + res_cost
             dem_txt = f"${dem_cost:,.2f} mxn"

        # 4. Mixed Use
        mix = calculate_mixed_use(usos_mixtos, num_locales, reg['cos_area'], costo_local_m2)

        # 5. Parking
        area_circulacion = reg['cus_area'] * areaCirculacionPorcentaje
        park = calculate_parking(
            estacionamiento, n_viviendas, reg['cos_area'], area_circulacion,
            delegacion, Distrito, tipo_estacionamiento, params
        )

        # 6. Costs
        # Correction: logic.py previously used 'cos_area' (footprint) which drastically underestimated cost.
        # Fixed to use 'cus_area' (total permitted area) and restored the original NoNA.py logic 
        # which sums construction cost twice (likely Base + Finishes/Indirects factor in original logic).
        
        base_construction = reg['cus_area'] * costoMetroConstruccion
        costos_directos = base_construction + dem_cost
        
        # Parking area usually does not consume CUS, but circulation does.
        area_venta = reg['cus_area'] - mix['area_comercio'] - area_circulacion
        
        ingreso_vivienda = area_venta * Costo_de_venta_m2
        ingreso_bruto_inicial = ingreso_vivienda + mix['ingreso_total']
        
        # Indirects
        # Indirects Parameters
        pct_honorarios = params.get('PCT_HONORARIOS', DEFAULT_PARAMS['PCT_HONORARIOS']) / 100.0
        pct_legales = params.get('PCT_LEGALES', DEFAULT_PARAMS['PCT_LEGALES']) / 100.0
        pct_adm = params.get('PCT_ADM', DEFAULT_PARAMS['PCT_ADM']) / 100.0
        pct_fin = params.get('PCT_FIN', DEFAULT_PARAMS['PCT_FIN']) / 100.0
        pct_com = params.get('PCT_COM', DEFAULT_PARAMS['PCT_COM']) / 100.0
        
        # Taxes
        iva_percent = float(data.get('iva_percent', 0.16))
        
        # EXPERT FIX: Isolating Fixed vs Variable Costs
        fixed_indirects = costos_directos * pct_honorarios
        var_indirects_inicial = ingreso_bruto_inicial * (pct_legales + pct_adm + pct_fin + pct_com)
        costos_indirectos = fixed_indirects + var_indirects_inicial
        
        base_construction_total = costos_directos + costos_indirectos + park['cost']
        monto_iva = base_construction_total * iva_percent
        costo_total_construccion = total_val + base_construction_total + monto_iva

        # 7. Simulation
        ganancia_bruta = ingreso_bruto_inicial - costo_total_construccion
        utilidad_actual = (ganancia_bruta / ingreso_bruto_inicial * 100.0) if ingreso_bruto_inicial > 0 else 0.0
        
        # EXPERT FINANCIAL ALGORITHM: Correct Circular Variable Cost Dependencies
        # We always simulate the required revenue to hit 'utilidadDeseada' to provide smart feedback.
        fixed_costs_before_iva = costos_directos + park['cost'] + fixed_indirects
        fixed_cost_total = total_val + (fixed_costs_before_iva * (1.0 + iva_percent))
        var_rate = (pct_legales + pct_adm + pct_fin + pct_com) * (1.0 + iva_percent)
        desired_margin = utilidadDeseada / 100.0
        
        target_rev = ingreso_bruto_inicial
        target_gain = ganancia_bruta
        target_util = utilidad_actual
        
        # Calculate optimal values
        if (1.0 - var_rate - desired_margin) > 0:
            target_rev = fixed_cost_total / (1.0 - var_rate - desired_margin)
            
            # Recalculate true P&L based on new target_rev for the optimized scenario
            var_indirects_final = target_rev * (pct_legales + pct_adm + pct_fin + pct_com)
            costos_indirectos_optimizado = fixed_indirects + var_indirects_final
            base_construction_total_optimizado = costos_directos + costos_indirectos_optimizado + park['cost']
            monto_iva_optimizado = base_construction_total_optimizado * iva_percent
            costo_total_construccion_optimizado = total_val + base_construction_total_optimizado + monto_iva_optimizado
            
            target_gain = target_rev - costo_total_construccion_optimizado
            target_util = (target_gain / target_rev * 100.0) if target_rev > 0 else 0.0
        else:
            costo_total_construccion_optimizado = costo_total_construccion

        # AI Recommendation values
        target_ingreso_vivienda = target_rev - mix['ingreso_total']
        target_precio_venta_m2 = target_ingreso_vivienda / area_venta if area_venta > 0 else Costo_de_venta_m2
        
        # New Metric: Cost per Apartment
        # Pro-rata total project cost based on residential vs commercial saleable area
        total_saleable = area_venta + mix['area_comercio']
        residential_cost_fraction = (area_venta / total_saleable) if total_saleable > 0 else 1.0
        costo_vivienda_total = costo_total_construccion * residential_cost_fraction
        
        costo_por_departamento = costo_vivienda_total / n_viviendas if n_viviendas > 0 else 0.0

        # --- EXPERT FIX: Parametric Catalog "Rayos X" ---
        # Base cost of just pure structure and interior architecture without indirects (Base Construction + Parking Base Cost)
        total_obra_pura = base_construction + park['cost']
        
        # We also distribute this between Vivienda vs Comercio
        obra_vivienda = total_obra_pura * residential_cost_fraction
        obra_comercial = total_obra_pura * (1.0 - residential_cost_fraction)
        
        # Standard Parametric Concept weights (Sum = 100%)
        # Note: Demolition is handled physically elsewhere, here we just split the fresh superstructure build costs.
        catalogo_pesos = [
            {"key": "preliminares", "name": "Preliminares y Terracerías", "peso": 0.02},
            {"key": "cimentacion", "name": "Cimentación", "peso": 0.08},
            {"key": "estructura", "name": "Estructura y Costillas", "peso": 0.25},
            {"key": "albanileria", "name": "Albañilería y Muros", "peso": 0.12},
            {"key": "hidrosanitaria", "name": "Instalaciones Hidrosanitarias", "peso": 0.07},
            {"key": "electrica", "name": "Instalaciones Eléctricas", "peso": 0.09},
            {"key": "hvac_especiales", "name": "HVAC y Sist. Especiales", "peso": 0.11},
            {"key": "acabados", "name": "Acabados y Recubrimientos", "peso": 0.14},
            {"key": "canceleria", "name": "Cancelería y Vidrio", "peso": 0.06},
            {"key": "carpinteria", "name": "Carpintería y Equipamiento", "peso": 0.05},
            {"key": "limpieza", "name": "Entregas y Herrería", "peso": 0.01},
        ]
        
        catalogo_obra = []
        for idx, partida in enumerate(catalogo_pesos):
            catalogo_obra.append({
                "clave": f"OBR-{str(idx+1).zfill(2)}",
                "concepto": partida['name'],
                "peso_porcentaje": partida['peso'] * 100.0,
                "monto_total": total_obra_pura * partida['peso'],
                "monto_vivienda": obra_vivienda * partida['peso'],
                "monto_comercial": obra_comercial * partida['peso']
            })

        # ═══════════════════════════════════════════════════════════════════════
        # CFO MODULE 11: INSTITUTIONAL-GRADE COMMERCIAL ANALYSIS
        # Venta (Especulativo) vs Renta (Patrimonial) — 24-Month Projection
        # ═══════════════════════════════════════════════════════════════════════

        cap_rate_anual = 0.09          # 9% Tasa de capitalización
        vacancia_estructural = 0.05    # 5% vacancy allowance
        mantenimiento_pct = 0.03       # 3% annual maintenance on commercial cost
        isr_renta_pct = 0.30           # 30% ISR on rental income
        comision_venta_pct = 0.04      # 4% brokerage commission on sale
        incremento_renta_anual = 0.04  # 4% annual rent escalation (inflation)

        # --- Allocated Commercial Investment ---
        costo_comercial_total = costo_total_construccion * (1.0 - residential_cost_fraction)
        ingreso_comercial_venta = mix['ingreso_total']
        num_locales_val = max(num_locales, 1)

        # --- Per-Local Metrics ---
        area_por_local = mix['area_local'] if mix['area_local'] > 0 else 0
        costo_por_local = costo_comercial_total / num_locales_val
        precio_venta_por_local = mix['venta_local']

        # --- Rental Pricing (Cap Rate Method) ---
        renta_anual_estimada = ingreso_comercial_venta * cap_rate_anual
        renta_mensual_estimada = renta_anual_estimada / 12.0
        precio_renta_m2 = (renta_mensual_estimada / mix['area_comercio']) if mix['area_comercio'] > 0 else 0
        renta_por_local = renta_mensual_estimada / num_locales_val

        # --- Yield on Cost (institutional metric) ---
        yield_on_cost = (renta_anual_estimada / costo_comercial_total * 100) if costo_comercial_total > 0 else 0

        # --- Net Operating Income (NOI) ---
        mantenimiento_anual = costo_comercial_total * mantenimiento_pct
        vacancia_anual = renta_anual_estimada * vacancia_estructural
        noi_anual = renta_anual_estimada - vacancia_anual - mantenimiento_anual
        noi_mensual = noi_anual / 12.0
        ingreso_neto_renta_mensual = noi_mensual * (1.0 - isr_renta_pct)  # After tax

        # --- Comision de Venta ---
        comision_venta_total = ingreso_comercial_venta * comision_venta_pct
        ingreso_neto_venta = ingreso_comercial_venta - comision_venta_total
        ganancia_venta = ingreso_neto_venta - costo_comercial_total
        margen_venta = (ganancia_venta / ingreso_neto_venta * 100) if ingreso_neto_venta > 0 else 0

        # --- Payback ---
        payback_renta_meses = (costo_comercial_total / ingreso_neto_renta_mensual) if ingreso_neto_renta_mensual > 0 else 0

        # --- Multi-Horizon ROI (Renta) ---
        roi_1y = ((noi_anual * (1 - isr_renta_pct)) / costo_comercial_total * 100) if costo_comercial_total > 0 else 0
        roi_2y = roi_1y * 2 * 1.04  # with escalation
        roi_5y = 0
        ingreso_acum_5y = 0
        for yr in range(5):
            escala = (1 + incremento_renta_anual) ** yr
            noi_yr = (renta_anual_estimada * escala * (1 - vacancia_estructural) - mantenimiento_anual) * (1 - isr_renta_pct)
            ingreso_acum_5y += noi_yr
        roi_5y = (ingreso_acum_5y / costo_comercial_total * 100) if costo_comercial_total > 0 else 0

        # ── TIME SERIES (24 months) ──────────────────────────────────
        meses_proyeccion = 24
        meses_obra = 12
        flujo_especulativo = []
        flujo_patrimonial = []
        acum_especulativo = 0.0
        acum_patrimonial = 0.0
        breakeven_venta_mes = 0
        breakeven_renta_mes = 0

        for mes in range(1, meses_proyeccion + 1):
            # ── EGRESOS (both scenarios share construction cost) ──
            egreso_construccion = -(costo_comercial_total / meses_obra) if mes <= meses_obra else 0
            egreso_mantenimiento = -(mantenimiento_anual / 12.0) if mes > meses_obra else 0

            # ═══ SCENARIO A: SELL ═══
            # Sales Curve: Enganche (20%) at pre-sale, Parcialidad during obra, Finiquito at delivery
            ingreso_venta = 0.0
            if mes == 4:     ingreso_venta = ingreso_comercial_venta * 0.10   # Pre-sale 10%
            elif mes == 6:   ingreso_venta = ingreso_comercial_venta * 0.10   # Enganche 10%
            elif mes == 8:   ingreso_venta = ingreso_comercial_venta * 0.10   # Parcialidad 10%
            elif mes == 10:  ingreso_venta = ingreso_comercial_venta * 0.10   # Parcialidad 10%
            elif mes == 12:  ingreso_venta = ingreso_comercial_venta * 0.15   # Pre-entrega 15%
            elif mes == 13:  ingreso_venta = ingreso_comercial_venta * 0.20   # Entrega 20%
            elif mes == 14:  ingreso_venta = ingreso_comercial_venta * 0.15   # Finiquito 15%
            elif mes == 16:  ingreso_venta = ingreso_comercial_venta * 0.10   # Rezago 10%

            egreso_comision_v = -(ingreso_venta * comision_venta_pct) if ingreso_venta > 0 else 0
            flujo_neto_v = egreso_construccion + ingreso_venta + egreso_comision_v
            acum_especulativo += flujo_neto_v

            if breakeven_venta_mes == 0 and acum_especulativo >= 0 and mes > 1:
                breakeven_venta_mes = mes

            flujo_especulativo.append({
                "mes": mes,
                "egreso": egreso_construccion + egreso_comision_v,
                "ingreso": ingreso_venta,
                "flujo_neto": flujo_neto_v,
                "acumulado": acum_especulativo
            })

            # ═══ SCENARIO B: RENT ═══
            ocupacion = 0.0
            if mes == 13:   ocupacion = 0.25
            elif mes == 14: ocupacion = 0.50
            elif mes == 15: ocupacion = 0.75
            elif mes >= 16: ocupacion = 1.0 - vacancia_estructural

            # Apply annual escalation after year 1 of renting
            meses_rentando = max(0, mes - 12)
            escala_renta = (1 + incremento_renta_anual) ** (meses_rentando / 12.0) if meses_rentando > 0 else 1.0
            renta_bruta = renta_mensual_estimada * escala_renta * ocupacion
            isr_mes = renta_bruta * isr_renta_pct if renta_bruta > 0 else 0
            ingreso_renta_neto = renta_bruta - isr_mes

            flujo_neto_r = egreso_construccion + egreso_mantenimiento + ingreso_renta_neto
            acum_patrimonial += flujo_neto_r

            if breakeven_renta_mes == 0 and acum_patrimonial >= 0 and mes > 1:
                breakeven_renta_mes = mes

            flujo_patrimonial.append({
                "mes": mes,
                "egreso": egreso_construccion + egreso_mantenimiento,
                "ingreso": ingreso_renta_neto,
                "renta_bruta": renta_bruta,
                "isr": isr_mes,
                "ocupacion": ocupacion,
                "flujo_neto": flujo_neto_r,
                "acumulado": acum_patrimonial
            })

        # ── Summary object for frontend ──
        cfo_summary = {
            "inversion_comercial": costo_comercial_total,
            "num_locales": num_locales_val,
            "area_por_local": area_por_local,
            "costo_por_local": costo_por_local,
            "precio_venta_local": precio_venta_por_local,
            "precio_venta_m2": costo_local_m2,
            "renta_por_local": renta_por_local,
            "renta_m2_mensual": precio_renta_m2,
            "cap_rate": cap_rate_anual,
            "yield_on_cost": yield_on_cost,
            "vacancia_pct": vacancia_estructural,
            "mantenimiento_pct": mantenimiento_pct,
            "isr_pct": isr_renta_pct,
            "comision_venta_pct": comision_venta_pct,
            "incremento_anual": incremento_renta_anual,
            "noi_anual": noi_anual,
            "noi_mensual": noi_mensual,
            "ingreso_neto_renta_mensual": ingreso_neto_renta_mensual,
            "renta_anual_bruta": renta_anual_estimada,
            "ingreso_neto_venta": ingreso_neto_venta,
            "ganancia_venta": ganancia_venta,
            "margen_venta": margen_venta,
            "comision_venta": comision_venta_total,
            "payback_meses": payback_renta_meses,
            "breakeven_venta_mes": breakeven_venta_mes,
            "breakeven_renta_mes": breakeven_renta_mes,
            "roi_1y": roi_1y,
            "roi_2y": roi_2y,
            "roi_5y": roi_5y,
            "acumulado_venta_final": acum_especulativo,
            "acumulado_renta_final": acum_patrimonial,
        }

        # ═══════════════════════════════════════════════════════════════════════
        # NONA INTELLIGENCE ENGINE v2.0
        # ═══════════════════════════════════════════════════════════════════════

        # ── Key ratios for analysis ──
        eficiencia_arq = (area_venta / reg['cus_area'] * 100) if reg['cus_area'] > 0 else 0
        ratio_terreno_inversion = (total_val / costo_total_construccion * 100) if costo_total_construccion > 0 else 0
        costo_ingreso_ratio = (costo_total_construccion / target_rev * 100) if target_rev > 0 else 100
        costo_m2_vendible = costo_total_construccion / area_venta if area_venta > 0 else 0
        ingreso_m2_vendible = target_rev / area_venta if area_venta > 0 else 0
        circ_pct = areaCirculacionPorcentaje * 100

        # ── Banking Metrics ──
        # DSCR: assuming 70% LTV, 12% annual rate, 15yr amortization
        ltv_assumed = 0.70
        tasa_anual_credito = 0.12
        plazo_anios = 15
        monto_credito = costo_total_construccion * ltv_assumed
        pago_mensual_credito = (monto_credito * (tasa_anual_credito / 12)) / (1 - (1 + tasa_anual_credito / 12) ** (-plazo_anios * 12)) if monto_credito > 0 else 0
        pago_anual_credito = pago_mensual_credito * 12
        # NOI for DSCR uses the TOTAL project income, not just commercial
        noi_total_proyecto = target_rev * (1.0 - var_rate) - costos_directos - fixed_indirects
        dscr = noi_total_proyecto / pago_anual_credito if pago_anual_credito > 0 else 0
        ltv_ratio = (monto_credito / target_rev * 100) if target_rev > 0 else 0

        banking_metrics = {
            "dscr": round(dscr, 2),
            "ltv_pct": round(ltv_ratio, 1),
            "monto_credito": monto_credito,
            "pago_mensual_credito": pago_mensual_credito,
            "costo_m2_vendible": costo_m2_vendible,
            "ingreso_m2_vendible": ingreso_m2_vendible,
            "costo_ingreso_ratio": round(costo_ingreso_ratio, 1),
            "ratio_terreno_inversion": round(ratio_terreno_inversion, 1),
            "eficiencia_arq": round(eficiencia_arq, 1),
        }

        # ── INSIGHTS ENGINE ──
        insights = []

        # Rule 1: Utility below target
        if utilidad_actual < utilidadDeseada - 0.5:
            gap = utilidadDeseada - utilidad_actual
            insights.append({
                "tipo": "warning",
                "icono": "trending-down",
                "titulo": "Utilidad por debajo de la meta",
                "descripcion": f"Tu utilidad real es del {utilidad_actual:.1f}%, {gap:.1f} puntos por debajo de tu meta del {utilidadDeseada:.1f}%. NoNA recomienda incrementar el precio de venta a ${target_precio_venta_m2:,.0f}/m² para alcanzar la viabilidad.",
                "accion": "auto_ajustar_precio"
            })

        # Rule 2: Utility above target (great news!)
        if utilidad_actual >= utilidadDeseada + 2:
            insights.append({
                "tipo": "success",
                "icono": "check-circle",
                "titulo": "¡Proyecto altamente rentable!",
                "descripcion": f"Tu utilidad del {utilidad_actual:.1f}% supera tu meta del {utilidadDeseada:.1f}% en {(utilidad_actual - utilidadDeseada):.1f} puntos. Tienes margen para absorber imprevistos o mejorar acabados.",
                "accion": None
            })

        # Rule 3: Renta is better than Venta
        if usos_mixtos and margen_venta < 15 and roi_5y > 25:
            insights.append({
                "tipo": "insight",
                "icono": "building",
                "titulo": "La renta patrimonial es la mejor estrategia",
                "descripcion": f"El margen de venta de locales es solo del {margen_venta:.1f}%, pero la renta genera un ROI del {roi_5y:.1f}% a 5 años con flujo perpetuo de ${ingreso_neto_renta_mensual:,.0f}/mes. NoNA recomienda retener los locales como patrimonio.",
                "accion": None
            })

        # Rule 4: Venta is better than Renta
        if usos_mixtos and margen_venta > 15 and payback_renta_meses > 36:
            insights.append({
                "tipo": "insight",
                "icono": "store",
                "titulo": "La venta especulativa maximiza tu retorno",
                "descripcion": f"Con un margen del {margen_venta:.1f}% y ganancia neta de ${ganancia_venta:,.0f}, la venta de locales genera liquidez inmediata. La renta requeriría {math.ceil(payback_renta_meses)} meses para igualar este retorno.",
                "accion": None
            })

        # Rule 5: Hybrid recommendation
        if usos_mixtos and 10 <= margen_venta <= 20 and roi_5y > 15 and num_locales >= 4:
            mitad = num_locales // 2
            insights.append({
                "tipo": "insight",
                "icono": "git-branch",
                "titulo": "Estrategia híbrida recomendada",
                "descripcion": f"Con {num_locales} locales y métricas equilibradas, considera vender {mitad} locales para recuperar capital y rentar {num_locales - mitad} para ingreso recurrente. Esto optimiza liquidez y patrimonio simultáneamente.",
                "accion": None
            })

        # Rule 6: Efficiency too low
        if eficiencia_arq < 65 and reg['cus_area'] > 0:
            area_perdida = reg['cus_area'] - area_venta
            insights.append({
                "tipo": "warning",
                "icono": "layout",
                "titulo": f"Eficiencia arquitectónica baja ({eficiencia_arq:.0f}%)",
                "descripcion": f"Estás perdiendo {area_perdida:.0f} m² en circulaciones, estacionamiento y áreas comunes. El estándar del mercado es >70%. Optimizar el layout generaría hasta ${(area_perdida * 0.3 * Costo_de_venta_m2):,.0f} en ingresos adicionales.",
                "accion": None
            })

        # Rule 7: Circulation too high
        if circ_pct > 22:
            reduction_area = reg['cus_area'] * (circ_pct - 20) / 100
            insights.append({
                "tipo": "tip",
                "icono": "move",
                "titulo": f"Circulación excesiva ({circ_pct:.0f}%)",
                "descripcion": f"Tu área de circulación supera el estándar del mercado (18-22%). Reducirla al 20% liberaría ~{reduction_area:.0f} m² vendibles adicionales, equivalentes a ~${(reduction_area * Costo_de_venta_m2):,.0f} en ingresos.",
                "accion": None
            })

        # Rule 8: Land too expensive
        if ratio_terreno_inversion > 35:
            insights.append({
                "tipo": "warning",
                "icono": "map-pin",
                "titulo": f"Terreno representa el {ratio_terreno_inversion:.0f}% de la inversión",
                "descripcion": f"Los proyectos eficientes mantienen el costo del terreno entre 20-30% de la inversión total. Este ratio elevado limita tu margen inherentemente. Considera negociar el precio o buscar un terreno más competitivo.",
                "accion": None
            })

        # Rule 9: Parking too expensive
        parking_ratio = (park['cost'] / costo_total_construccion * 100) if costo_total_construccion > 0 else 0
        if parking_ratio > 15:
            insights.append({
                "tipo": "tip",
                "icono": "car",
                "titulo": f"Estacionamiento costoso ({parking_ratio:.0f}% del proyecto)",
                "descripcion": f"El estacionamiento cuesta ${park['cost']:,.0f}, representando el {parking_ratio:.0f}% del costo total. Considera un esquema semi-subterráneo o mecánico para reducir hasta un 30% este costo.",
                "accion": None
            })

        # Rule 10: DSCR warning
        if dscr > 0 and dscr < 1.25:
            insights.append({
                "tipo": "warning",
                "icono": "shield-alert",
                "titulo": f"Cobertura de deuda ajustada (DSCR: {dscr:.2f}x)",
                "descripcion": f"Un DSCR menor a 1.25x puede dificultar la obtención de crédito bancario. Los bancos generalmente requieren DSCR > 1.30x. Incrementar ingresos o reducir costos mejoraría esta métrica.",
                "accion": None
            })
        elif dscr >= 1.5:
            insights.append({
                "tipo": "success",
                "icono": "shield-check",
                "titulo": f"Excelente cobertura de deuda (DSCR: {dscr:.2f}x)",
                "descripcion": f"Un DSCR de {dscr:.2f}x está muy por encima del mínimo bancario (1.25x). Este proyecto es altamente financiable y atractivo para instituciones crediticias.",
                "accion": None
            })

        # Rule 11: Add more locales suggestion
        if usos_mixtos and num_locales > 0 and num_locales <= 4 and margen_venta > 5:
            extra = 2
            nueva_renta = renta_mensual_estimada * (num_locales + extra) / num_locales
            insights.append({
                "tipo": "tip",
                "icono": "plus-circle",
                "titulo": "Potencial de más locales comerciales",
                "descripcion": f"Si agregaras {extra} locales más (de {num_locales} a {num_locales + extra}), la renta mensual estimada subiría de ${renta_mensual_estimada:,.0f} a ${nueva_renta:,.0f}/mes, fortaleciendo significativamente el modelo patrimonial.",
                "accion": None
            })

        # Rule 12: Cost per apartment
        if n_viviendas > 0 and costo_por_departamento > 0:
            precio_promedio_depto = ingreso_vivienda / n_viviendas if n_viviendas > 0 else 0
            margen_depto = ((precio_promedio_depto - costo_por_departamento) / precio_promedio_depto * 100) if precio_promedio_depto > 0 else 0
            if margen_depto < 15:
                insights.append({
                    "tipo": "warning",
                    "icono": "home",
                    "titulo": f"Margen por departamento bajo ({margen_depto:.0f}%)",
                    "descripcion": f"Cada departamento cuesta ${costo_por_departamento:,.0f} y se vende a ${precio_promedio_depto:,.0f}, dejando solo {margen_depto:.1f}% de margen. Evalúa subir precio de venta o reducir costo de construcción.",
                    "accion": None
                })

        # Rule 13: Great cost-income ratio
        if costo_ingreso_ratio < 70:
            insights.append({
                "tipo": "success",
                "icono": "zap",
                "titulo": f"Relación costo/ingreso excelente ({costo_ingreso_ratio:.0f}%)",
                "descripcion": f"Solo el {costo_ingreso_ratio:.0f}% de tus ingresos se consume en costos totales. Esto está por encima del benchmark del mercado (normalmente 75-85%), indicando un proyecto altamente eficiente.",
                "accion": None
            })

        # ── COMMERCIAL RECOMMENDATION ──
        comercial_recomendacion = "sin_locales"
        comercial_razon = ""
        if usos_mixtos and num_locales > 0:
            if margen_venta > 15 and payback_renta_meses > 36:
                comercial_recomendacion = "vender"
                comercial_razon = f"El margen de venta ({margen_venta:.1f}%) es sólido y el payback de renta ({math.ceil(payback_renta_meses)} meses) es demasiado largo. La venta especulativa maximiza tu retorno con liquidez inmediata."
            elif margen_venta < 10 and roi_5y > 25:
                comercial_recomendacion = "rentar"
                comercial_razon = f"El margen de venta es bajo ({margen_venta:.1f}%), pero la renta genera ROI de {roi_5y:.1f}% a 5 años con ingreso perpétuo de ${ingreso_neto_renta_mensual:,.0f}/mes post-ISR. El patrimonio es tu mejor estrategia."
            elif margen_venta >= 10 and margen_venta <= 20 and roi_5y > 15 and num_locales >= 3:
                comercial_recomendacion = "hibrida"
                mitad = num_locales // 2
                comercial_razon = f"Ambos escenarios son viables. Recomendamos vender {mitad} locales para liquidez inmediata y rentar {num_locales - mitad} para ingreso recurrente. Esto optimiza riesgo y retorno."
            elif margen_venta > 20:
                comercial_recomendacion = "vender"
                comercial_razon = f"Con un margen excepcional del {margen_venta:.1f}% y ganancia de ${ganancia_venta:,.0f}, la venta es claramente la opción óptima para este proyecto."
            elif ganancia_venta < 0:
                comercial_recomendacion = "rentar"
                comercial_razon = f"La venta de locales genera pérdida ({margen_venta:.1f}%). La renta patrimonial es la única estrategia viable, generando ${ingreso_neto_renta_mensual:,.0f}/mes netos."
            else:
                comercial_recomendacion = "evaluar"
                comercial_razon = f"Las métricas son mixtas (margen venta: {margen_venta:.1f}%, ROI renta 5Y: {roi_5y:.1f}%). Se recomienda sensibilizar el precio de venta de locales para tomar la decisión correcta."

        # ── VIABILITY SCORE (0-100) ──
        def score_dim(val, low, high):
            """Map value to 0-100 between low and high."""
            if val <= low: return 0
            if val >= high: return 100
            return ((val - low) / (high - low)) * 100

        dim_utilidad = score_dim(utilidad_actual, 0, utilidadDeseada * 1.2)
        dim_eficiencia = score_dim(eficiencia_arq, 40, 85)
        dim_costo_ingreso = score_dim(100 - costo_ingreso_ratio, 0, 40)  # inverted: lower ratio = better
        dim_cashflow = score_dim(breakeven_venta_mes, 24, 6) if breakeven_venta_mes > 0 else 30  # earlier = better
        dim_cashflow = min(100, max(0, 100 - dim_cashflow + 50)) if breakeven_venta_mes > 0 else 40
        dim_dscr = score_dim(dscr, 0.8, 2.0)

        # Weighted average
        score_total = (
            dim_utilidad * 0.30 +
            dim_eficiencia * 0.20 +
            dim_costo_ingreso * 0.20 +
            dim_cashflow * 0.15 +
            dim_dscr * 0.15
        )

        viability_score = {
            "total": round(min(100, max(0, score_total))),
            "dimensiones": {
                "utilidad": round(min(100, max(0, dim_utilidad))),
                "eficiencia": round(min(100, max(0, dim_eficiencia))),
                "costo_ingreso": round(min(100, max(0, dim_costo_ingreso))),
                "cashflow": round(min(100, max(0, dim_cashflow))),
                "financiabilidad": round(min(100, max(0, dim_dscr))),
            },
            "label": "Excelente" if score_total >= 80 else "Bueno" if score_total >= 60 else "Regular" if score_total >= 40 else "Riesgoso"
        }

        # ── SCORE OPTIMIZER ──────────────────────────────────
        # When score < 80, calculate exactly what changes push it above 80
        optimizer = None
        if score_total < 80:
            sugerencias = []
            target_score = 82  # aim slightly above 80 for margin
            deficit = target_score - score_total

            dims = viability_score["dimensiones"]
            weights = {"utilidad": 0.30, "eficiencia": 0.20, "costo_ingreso": 0.20, "cashflow": 0.15, "financiabilidad": 0.15}

            # Sort dimensions by how much room for improvement they have (lowest first)
            ranked = sorted(dims.items(), key=lambda x: x[1])

            # For each weak dimension, calculate what real-world parameter fixes it
            for dim_name, dim_val in ranked:
                if dim_val >= 85:
                    continue  # already strong, skip this one

                # How much does this dimension need to improve to help reach 80?
                # We want to bring weak dims to at least 80
                dim_target = max(80, dim_val + (deficit / weights[dim_name]) * 0.5)
                dim_target = min(100, dim_target)
                improvement_needed = dim_target - dim_val

                if improvement_needed < 5:
                    continue

                if dim_name == "utilidad" and area_venta > 0:
                    # To reach dim_target utilidad, we need utilidad_actual to be:
                    # score_dim(utilidad_actual, 0, utilidadDeseada * 1.2) = dim_target
                    # → utilidad_actual = dim_target/100 * utilidadDeseada * 1.2
                    target_util_val = (dim_target / 100.0) * utilidadDeseada * 1.2
                    # Price needed: target_rev = fixed_cost_total / (1 - var_rate - target_util_val/100)
                    denom = (1.0 - var_rate - target_util_val / 100.0)
                    if denom > 0.01:
                        new_target_rev = fixed_cost_total / denom
                        new_ingreso_viv = new_target_rev - mix['ingreso_total']
                        new_precio_m2 = new_ingreso_viv / area_venta if area_venta > 0 else Costo_de_venta_m2
                        if new_precio_m2 > Costo_de_venta_m2:
                            sugerencias.append({
                                "dimension": "Utilidad",
                                "parametro": "Costo_de_venta_m2",
                                "parametro_label": "Precio de Venta por m²",
                                "valor_actual": round(Costo_de_venta_m2, 2),
                                "valor_sugerido": round(new_precio_m2, 2),
                                "impacto": f"Utilidad sube de {utilidad_actual:.1f}% a ~{target_util_val:.1f}%",
                                "dim_actual": dim_val,
                                "dim_nueva": round(dim_target),
                            })

                elif dim_name == "eficiencia" and reg['cus_area'] > 0:
                    # score_dim(eficiencia_arq, 40, 85) = dim_target
                    # eficiencia_arq = dim_target/100 * (85-40) + 40
                    target_eff = (dim_target / 100.0) * (85 - 40) + 40
                    # eficiencia = area_venta / cus_area * 100
                    # → new_circ_pct: area_venta = cus_area * (1 - new_circ_pct) - area_comercio...
                    # Simpler: target_circ_pct = 1 - (target_eff/100 * cus_area + mix['area_comercio']) / cus_area
                    needed_venta = target_eff / 100.0 * reg['cus_area']
                    if needed_venta > area_venta:
                        # Reduce circulation to gain more sellable area
                        new_circ_pct = max(0.10, (reg['cus_area'] - needed_venta - mix['area_comercio']) / reg['cus_area'])
                        if new_circ_pct < areaCirculacionPorcentaje:
                            sugerencias.append({
                                "dimension": "Eficiencia",
                                "parametro": "areaCirculacionPorcentaje",
                                "parametro_label": "Área de Circulación",
                                "valor_actual": round(areaCirculacionPorcentaje * 100, 1),
                                "valor_sugerido": round(new_circ_pct * 100, 1),
                                "impacto": f"Eficiencia sube de {eficiencia_arq:.0f}% a ~{target_eff:.0f}%",
                                "dim_actual": dim_val,
                                "dim_nueva": round(dim_target),
                                "unidad": "%"
                            })

                elif dim_name == "costo_ingreso":
                    # score_dim(100 - costo_ingreso_ratio, 0, 40) = dim_target
                    # → 100 - costo_ingreso_ratio = dim_target/100 * 40
                    target_ratio = 100 - (dim_target / 100.0 * 40)
                    # costo_ingreso_ratio = costo_total / target_rev * 100
                    # → to lower this: either increase revenue or decrease cost
                    # Option: decrease construction cost
                    if target_rev > 0:
                        new_costo_total = target_ratio / 100.0 * target_rev
                        reduccion = costo_total_construccion - new_costo_total
                        if reduccion > 0 and base_construction > 0 and reg['cus_area'] > 0:
                            # How much to reduce costoMetroConstruccion?
                            new_costo_m2 = costoMetroConstruccion - (reduccion / reg['cus_area'])
                            if new_costo_m2 > 0 and new_costo_m2 < costoMetroConstruccion:
                                sugerencias.append({
                                    "dimension": "Costo/Ingreso",
                                    "parametro": "costoMetroConstruccion",
                                    "parametro_label": "Costo de Construcción por m²",
                                    "valor_actual": round(costoMetroConstruccion, 2),
                                    "valor_sugerido": round(new_costo_m2, 2),
                                    "impacto": f"Ratio baja de {costo_ingreso_ratio:.0f}% a ~{target_ratio:.0f}%",
                                    "dim_actual": dim_val,
                                    "dim_nueva": round(dim_target),
                                })

            optimizer = {
                "necesario": score_total < 80,
                "score_actual": round(score_total),
                "score_objetivo": target_score,
                "sugerencias": sugerencias[:4],  # max 4 suggestions
            }

        viability_score["optimizer"] = optimizer

        return {
            "metrics": {
                # --- 1. LAND ---
                "Text_Area_Terreno": txt_area,
                "Text_Valor_Terreno": txt_val,
                "Text_Costo_Unitario_Tierra": f"${valor_terreno:,.2f} mxn", # New

                # --- 2. NORMATIVE ---
                "Text_COS_Area": f"{reg['cos_area']:.2f} m2",
                "Text_CUS_Area": f"{reg['cus_area']:.2f} m2",
                "Text_CAS_Area": f"{reg['cas_area']:.2f} m2",
                "Text_Net_Area": f"{reg['net_area']:.2f} m2",

                # --- 3. DEMOLITION ---
                "Text_Demolicion_Costo": f"${dem_cost_only:,.2f}",
                "Text_Licencia_Costo": f"${lic_cost:,.2f}",
                "Text_Residuos_Costo": f"${res_cost:,.2f}",
                "Text_Total_Demolicion": dem_txt,

                # --- 4. PARKING ---
                "Text_Cajones_Vivienda": f"{sum(park['details']['cajones_vivienda']):.1f}", 
                "Text_Cajones_Comercio": f"{sum(park['details']['cajones_comercio']):.1f}",
                "Text_Cajones_Total": f"{sum(park['details']['cajones_total']):.0f}",
                "Text_Area_Estacionamiento": f"{park['area']:.2f} m2",
                "Text_Costo_Estacionamiento": f"${park['cost']:,.2f}",

                # --- 5. AREAS ---
                "Text_Area_Circulacion": f"{area_circulacion:.2f} m2",
                "Text_Area_Comercio": f"{mix['area_comercio']:.2f} m2",
                "Text_Area_Vendible_Vivienda": f"{area_venta:.2f} m2",
                "Text_Eficiencia": f"{((area_venta / reg['cus_area']) * 100):.2f}%" if reg['cus_area'] else "0%",

                # --- 6. COST ANALYSIS ---
                "Text_Construccion_Base": f"${base_construction:,.2f}",
                "Text_Costos_Directos": f"${costos_directos:,.2f}",
                
                # Indirects Breakdown using real values
                "Text_Honorarios": f"${fixed_indirects:,.2f}",
                "Text_Legales": f"${(ingreso_bruto_inicial * pct_legales):,.2f}",
                "Text_Administrativos": f"${(ingreso_bruto_inicial * pct_adm):,.2f}",
                "Text_Financieros": f"${(ingreso_bruto_inicial * pct_fin):,.2f}",
                "Text_Comerciales": f"${(ingreso_bruto_inicial * pct_com):,.2f}",
                "Text_Costos_Indirectos": f"${costos_indirectos:,.2f}",
                
                "Text_Monto_IVA": f"${monto_iva:,.2f}",
                "Text_Costo_Total": f"${costo_total_construccion:,.2f}",

                # --- 7. INCOME ---
                "Text_Ingreso_Vivienda": f"${ingreso_vivienda:,.2f}",
                "Text_Ingreso_Locales": f"${mix['ingreso_total']:,.2f}",
                "Text_Ingreso_Total_Inicial": f"${ingreso_bruto_inicial:,.2f}",
                "Text_Ingreso_Total_Optimizado": f"${target_rev:,.2f}",
                "Text_Precio_Promedio_M2": f"${(ingreso_vivienda / area_venta):,.2f}" if area_venta else "$0",

                # --- 8. PROFITABILITY ---
                "Text_Utilidad_Inicial": f"{utilidad_actual:.2f}%",
                "Text_Utilidad_Final": f"{target_util:.2f}%",
                "Text_Ganancia_Bruta": f"${ganancia_bruta:,.2f}",
                "Text_Ganancia_Neta": f"${(ganancia_bruta * 0.70):,.2f}", # EXPERT FIX: Deduct 30% ISR
                "Text_ROI": f"{((ganancia_bruta / costo_total_construccion) * 100):.2f}%" if costo_total_construccion else "0%",
                
                # --- AI ADVISORY ---
                "Text_Target_Precio_Venta_M2": f"${target_precio_venta_m2:,.2f}",
                
                # --- 9. METRICS ---
                "Text_Costo_Por_Depto": f"${costo_por_departamento:,.2f}",
                "Text_Precio_Promedio_Vivienda": f"${(ingreso_vivienda / n_viviendas):,.2f}" if n_viviendas else "$0",
                "Text_Area_Promedio_Vivienda": f"{(area_venta / n_viviendas):.2f} m2" if n_viviendas else "0 m2",
                "Text_Punto_Equilibrio": f"${(costo_total_construccion / 0.7):,.2f}", # Simple break-even heuristic
                
                # --- 10. TIMELINE (Estimates) ---
                "Text_Meses_Tramites": "3 meses",
                "Text_Meses_Obra": "12 meses", # Generic placeholder
                "Text_Meses_Venta": "6 meses",
                "Text_Duracion_Total": "21 meses",
                
                # --- 11. CFO PATRIMONIAL ---
                "Text_Renta_Mensual_Sugerida": f"${renta_mensual_estimada:,.2f} / mes",
                "Text_Precio_Renta_M2": f"${precio_renta_m2:,.2f} / mes",
                "Text_Cap_Rate": f"{cap_rate_anual * 100:.1f}%",
                "Text_Payback_Meses": f"{math.ceil(payback_renta_meses)} meses" if payback_renta_meses > 0 else "N/A",
                "Text_Costo_Comercial": f"${costo_comercial_total:,.2f}"
            },
            "raw": {
                "area_terreno": area_val,
                "valor_terreno": total_val,
                
                # Normative
                "cos_area": reg['cos_area'],
                "cus_area": reg['cus_area'],
                "cas_area": reg['cas_area'],
                "net_area": reg['net_area'],
                
                # Demolition specifics
                "dem_cost_only": dem_cost_only,
                "lic_cost": lic_cost,
                "res_cost": res_cost,
                "total_dem_cost": dem_cost,

                # Areas
                "area_venta_vivienda": area_venta,
                "area_locales": mix['area_local'],
                "area_circulacion": area_circulacion,
                "area_comercio": mix['area_comercio'],
                
                # Costs
                "costo_directo": costos_directos,
                "base_construction": base_construction,
                "costo_indirecto": costos_indirectos,
                "costos_indirectos_desglose": {
                    "honorarios": fixed_indirects,
                    "legales": target_rev * pct_legales,
                    "administrativos": target_rev * pct_adm,
                    "financieros": target_rev * pct_fin,
                    "comerciales": target_rev * pct_com,
                },
                "costo_total": costo_total_construccion,
                "monto_iva": monto_iva,
                
                # Income
                "ingreso_inicial": ingreso_bruto_inicial,
                "ingreso_optimizado": target_rev,
                "ingreso_ventas_locales": mix['ingreso_total'],
                "ingreso_ventas_vivienda": target_rev - mix['ingreso_total'], 
                "target_precio_venta_m2": target_precio_venta_m2,
                
                # CFO module
                "flujo_especulativo": flujo_especulativo,
                "flujo_patrimonial": flujo_patrimonial,
                "cfo": cfo_summary,
                "renta_mensual_estimada": renta_mensual_estimada,
                "payback_renta_meses": payback_renta_meses,
                "costo_comercial_total": costo_comercial_total,
                "cap_rate": cap_rate_anual,
                
                # Profit
                "utilidad_inicial": utilidad_actual,
                "utilidad_optimizada": target_util, 
                "utilidad_monto": target_gain, 
                "roi": (target_gain / costo_total_construccion_optimizado * 100) if costo_total_construccion_optimizado else 0,
                
                # Parking
                "parking_cost": park['cost'],
                "parking_area": park['area'],
                "parking_spots": sum(park['details']['cajones_total']) if park['details'] else 0,
                "parking_spots_res": sum(park['details']['cajones_vivienda']) if park['details'] else 0,
                "parking_spots_com": sum(park['details']['cajones_comercio']) if park['details'] else 0,
                
                # Project
                "n_viviendas": n_viviendas,
                "costo_por_departamento": costo_por_departamento,
                "eficiencia": ((area_venta / reg['cus_area']) * 100) if reg['cus_area'] else 0,
                
                # Parametric X-Ray
                "catalogo_obra": catalogo_obra,
                
                # Intelligence Engine v2.0
                "insights": insights,
                "viability_score": viability_score,
                "banking": banking_metrics,
                "comercial_recomendacion": comercial_recomendacion,
                "comercial_razon": comercial_razon,
            }
        }
        
    except Exception as e:
        return {"error": str(e)}

import io
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel_content(result: Dict[str, Any]) -> bytes:
    """
    Generates a highly detailed, professional Excel workbook using openpyxl.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte NoNA"
    
    # 1. Disable Gridlines
    ws.sheet_view.showGridLines = False
    
    # Define Colors & Styles
    BRAND_BLUE = "2563EB"
    BRAND_SLATE = "0F172A"
    LIGHT_GRAY = "F8FAFC"
    MEDIUM_GRAY = "E2E8F0"
    GREEN_ACCENT = "10B981"
    
    header_fill = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_font = Font(color=BRAND_BLUE, bold=True, size=11)
    
    title_font = Font(color=BRAND_SLATE, bold=True, size=18)
    subtitle_font = Font(color=BRAND_BLUE, bold=True, size=14)
    normal_font = Font(color="475569", size=11)
    
    thin_border_bottom = Border(bottom=Side(style='thin', color=MEDIUM_GRAY))
    
    # Native Formats
    MONEY_FORMAT = '"$"#,##0.00_-'
    AREA_FORMAT = '#,##0.00 "m²"'
    PCT_FORMAT = '0.00%'
    
    # Setup Title Header
    ws.append(["NoNA", "I.Tech", "Reporte Financiero y Arquitectónico"])
    ws.append(["Generado automáticamente por NoNA Platform"])
    ws.append([])
    
    ws['A1'].font = title_font
    ws['B1'].font = subtitle_font
    ws['C1'].font = Font(color="64748B", italic=True, size=12)
    ws['A2'].font = Font(color="94A3B8", italic=True)
    
    metrics = result.get("metrics", {})
    raw = result.get("raw", {})
    
    if "error" in result:
        ws.append(["Error en Cálculo", result["error"]])
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
        
    ws.append(["Métrica", "Valor"])
    ws['A4'].fill = header_fill
    ws['A4'].font = header_font
    ws['B4'].fill = header_fill
    ws['B4'].font = header_font
    ws['A4'].alignment = Alignment(horizontal="center")
    ws['B4'].alignment = Alignment(horizontal="center")
    
    current_row = 5
    
    def add_section(title):
        nonlocal current_row
        ws.append(["", ""]) # spacer
        current_row += 1
        
        ws.append([title.upper(), ""])
        for col in ['A', 'B']:
            cell = ws[f'{col}{current_row}']
            cell.font = section_font
            cell.border = Border(bottom=Side(style='thick', color=BRAND_BLUE))
        current_row += 1
        
    def add_row(label, val, fmt=None, highlight=False):
        nonlocal current_row
        ws.append([label, val if val is not None else 0])
        
        cell_a = ws[f'A{current_row}']
        cell_b = ws[f'B{current_row}']
        
        cell_a.font = Font(bold=True, color=BRAND_SLATE if highlight else "334155")
        cell_b.font = Font(bold=highlight, color=GREEN_ACCENT if highlight else "475569")
        cell_b.alignment = Alignment(horizontal="right")
        
        # Zebra Striping
        if current_row % 2 == 0:
            fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
            cell_a.fill = fill
            cell_b.fill = fill
            
        cell_a.border = thin_border_bottom
        cell_b.border = thin_border_bottom
        
        if fmt:
            cell_b.number_format = fmt
        
        current_row += 1

    # Unpack raw values to use native numbers instead of formatted strings
    add_section("1. Resumen Ejecutivo (KPIs)")
    add_row("Utilidad Final Meta", raw.get("Text_Utilidad_Final_Raw", metrics.get("Text_Utilidad_Final")), fmt=PCT_FORMAT if isinstance(raw.get("Text_Utilidad_Final_Raw"), (int,float)) else None, highlight=True)
    add_row("Precio de Venta por Vv", raw.get("precio_promedio_vivienda", 0), fmt=MONEY_FORMAT)
    add_row("Costo Total del Proyecto", raw.get("costo_total", 0), fmt=MONEY_FORMAT)
    add_row("Área Vendible Total", raw.get("area_venta_vivienda", 0), fmt=AREA_FORMAT)
    add_row("Costo Cons. por Depto", raw.get("Text_Costo_Por_Depto_Raw", metrics.get("Text_Costo_Por_Depto")), fmt=MONEY_FORMAT if isinstance(raw.get("Text_Costo_Por_Depto_Raw"), (int,float)) else None)
    
    add_section("2. Terreno y Demolición")
    add_row("Área del Terreno", raw.get("area_terreno", 0), fmt=AREA_FORMAT)
    add_row("Valor Total del Terreno", raw.get("valor_terreno", 0), fmt=MONEY_FORMAT)
    add_row("Costo Total Demolición", raw.get("costo_total_demolicion", 0), fmt=MONEY_FORMAT)
    
    add_section("3. Normativa y Áreas")
    add_row("Área COS (Desplante)", raw.get("cos_area", 0), fmt=AREA_FORMAT)
    add_row("Área CUS (Cons. Max)", raw.get("cus_area", 0), fmt=AREA_FORMAT)
    add_row("Área CAS (Área Libre)", raw.get("cas_area", 0), fmt=AREA_FORMAT)
    
    add_section("4. Desglose de Costos de Construcción")
    add_row("Total Costos Directos", raw.get("costo_directo", 0), fmt=MONEY_FORMAT)
    add_row("Total Costos Indirectos", raw.get("costo_indirecto", 0), fmt=MONEY_FORMAT)
    add_row("Monto de IVA Estimado", raw.get("monto_iva", 0), fmt=MONEY_FORMAT)
    add_row("COSTO TOTAL DEL PROYECTO", raw.get("costo_total", 0), fmt=MONEY_FORMAT, highlight=True)
    
    add_section("5. Análisis Financiero")
    add_row("Ingreso por Ventas (Comercio)", raw.get("ingreso_ventas_locales", 0), fmt=MONEY_FORMAT)
    add_row("Ingreso por Ventas (Vivienda)", raw.get("ingreso_ventas_vivienda", 0), fmt=MONEY_FORMAT)
    add_row("Ingreso Total Optimizado", raw.get("ingreso_total_optimizado", 0), fmt=MONEY_FORMAT, highlight=True)
    add_row("Ganancia Neta", raw.get("utilidad_optimizada", 0), fmt=MONEY_FORMAT, highlight=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 28
    
    # ================= CHARTS =================
    
    # Write hidden data for charts (Cols Z, AA)
    hide_col_z = 26
    hide_col_aa = 27
    
    # Chart 1 Data: Bar Chart (General Analysis)
    bar_data = [
        ("Costo Total", raw.get("costo_total", 0)),
        ("Ingreso Meta", raw.get("ingreso_total_optimizado", 0)),
        ("Utilidad Neta", raw.get("utilidad_optimizada", 0))
    ]
    bar_start_row = 5
    for i, (cat, val) in enumerate(bar_data):
        ws.cell(row=bar_start_row+i, column=hide_col_z, value=cat)
        ws.cell(row=bar_start_row+i, column=hide_col_aa, value=val)
        
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 11
    bar_chart.title = "Balance Financiero General"
    bar_chart.y_axis.title = "Monto (MXN)"
    bar_chart.y_axis.number_format = '"$"#,##0'
    
    data_ref = Reference(ws, min_col=hide_col_aa, min_row=bar_start_row, max_row=bar_start_row+len(bar_data)-1)
    cats_ref = Reference(ws, min_col=hide_col_z, min_row=bar_start_row, max_row=bar_start_row+len(bar_data)-1)
    bar_chart.add_data(data_ref, titles_from_data=False)
    bar_chart.set_categories(cats_ref)
    bar_chart.shape = 4
    bar_chart.width = 16
    bar_chart.height = 8
    ws.add_chart(bar_chart, "D5")

    # Chart 2 Data: Pie Chart (Cost Structure)
    pie_data = [
        ("Tierra", raw.get("valor_terreno", 0)),
        ("Construcción", raw.get("costo_directo", 0) + raw.get("parking_cost", 0)),
        ("Indirectos e IVA", raw.get("costo_indirecto", 0) + raw.get("monto_iva", 0))
    ]
    pie_start_row = bar_start_row + len(bar_data) + 2
    for i, (cat, val) in enumerate(pie_data):
        ws.cell(row=pie_start_row+i, column=hide_col_z, value=cat)
        ws.cell(row=pie_start_row+i, column=hide_col_aa, value=val)
        
    pie_chart = PieChart()
    pie_chart.title = "Estructura de Costos del Proyecto"
    
    pdata_ref = Reference(ws, min_col=hide_col_aa, min_row=pie_start_row, max_row=pie_start_row+len(pie_data)-1)
    pcats_ref = Reference(ws, min_col=hide_col_z, min_row=pie_start_row, max_row=pie_start_row+len(pie_data)-1)
    pie_chart.add_data(pdata_ref, titles_from_data=False)
    pie_chart.set_categories(pcats_ref)
    pie_chart.width = 16
    pie_chart.height = 8
    ws.add_chart(pie_chart, "D20")
    
    # Hide the data columns mapping (Z, AA)
    ws.column_dimensions[openpyxl.utils.get_column_letter(hide_col_z)].hidden = True
    ws.column_dimensions[openpyxl.utils.get_column_letter(hide_col_aa)].hidden = True
    
    # ================= WORKSHEET 2: CÁTALOGO PARAMÉTRICO =================
    if "catalogo_obra" in raw:
        ws2 = wb.create_sheet(title="Catálogo Paramétrico")
        ws2.sheet_view.showGridLines = False
        
        ws2.append(["Catálogo Estructurado de Obra (Presupuesto Paramétrico)"])
        ws2['A1'].font = title_font
        ws2.append(["Distribución proyectada por especialidad sobre costo directo puro de edificación."])
        ws2['A2'].font = Font(color="94A3B8", italic=True)
        ws2.append([])
        
        headers = ["Clave", "Concepto de Obra", "Ponderación (%)", "Monto Total", "Incidencia Vivienda", "Incidencia Comercial"]
        ws2.append(headers)
        
        for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F']):
            cell = ws2[f'{col_letter}4']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        r_idx = 5
        for partida in raw["catalogo_obra"]:
            ws2.append([
                partida["clave"],
                partida["concepto"],
                partida["peso_porcentaje"] / 100.0, # Excel internal percentages
                partida["monto_total"],
                partida["monto_vivienda"],
                partida["monto_comercial"]
            ])
            
            # Formatting
            ws2[f'A{r_idx}'].font = Font(bold=True, color="64748B")
            ws2[f'B{r_idx}'].font = Font(bold=True, color="1E293B")
            ws2[f'C{r_idx}'].number_format = '0.00%'
            ws2[f'D{r_idx}'].number_format = MONEY_FORMAT
            ws2[f'E{r_idx}'].number_format = MONEY_FORMAT
            ws2[f'F{r_idx}'].number_format = MONEY_FORMAT
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws2[f'{col}{r_idx}']
                cell.border = thin_border_bottom
                if r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
            r_idx += 1
            
        # Totals Row
        ws2.append(["", "TOTAL COSTO DIRECTO", 1.0, raw.get("base_construction", 0) + raw.get("parking_cost", 0), "", ""])
        for col in ['B', 'C', 'D']:
            cell = ws2[f'{col}{r_idx}']
            cell.font = Font(bold=True, color=BRAND_BLUE)
            cell.border = Border(top=Side(style='thick', color=BRAND_BLUE))
        ws2[f'C{r_idx}'].number_format = '0.00%'
        ws2[f'D{r_idx}'].number_format = MONEY_FORMAT
        
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 22
        ws2.column_dimensions['E'].width = 22
        ws2.column_dimensions['F'].width = 22

    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
